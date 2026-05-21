NS = "ns"
import openpyxl
import json
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import range_boundaries
import re
from openpyxl.styles.colors import COLOR_INDEX, Color
from datetime import datetime
from copy import copy
from openpyxl.utils import get_column_letter

def apply_tint(rgb, tint):
    """
    Adjusts the RGB color according to tint value.
    Tint is a float from -1 (darken) to 1 (lighten).
    """
    if tint is None or tint == 0:
        return rgb
    import colorsys

    r, g, b = [x / 255.0 for x in rgb]
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1 + tint)
    else:
        l = l + (1 - l) * tint
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return (int(r * 255), int(g * 255), int(b * 255))


def rgb_from_argb(argb):
    """Convert ARGB (e.g. FF112233) hex string to RGB tuple."""
    a = argb[:2]
    r = int(argb[2:4], 16)
    g = int(argb[4:6], 16)
    b = int(argb[6:8], 16)
    return (r, g, b)


def rgb_to_hex(rgb):
    """Convert RGB tuple to ARGB uppercase hex string with alpha FF."""
    return 'FF' + ''.join(f'{v:02X}' for v in rgb)


def get_fill_color(cell):
    if not cell or not cell.fill:
        return 'FFFFFFFF'  # default white

    fill = cell.fill
    if fill.patternType is None or fill.patternType == 'none':
        return 'FFFFFFFF'  # no fill

    # Prefer fgColor but if pattern is not solid, may consider bgColor for actual color
    color = fill.fgColor
    if not color or (color.type == 'theme' and color.theme is None and not getattr(color, 'rgb', None)):
        color = fill.fgColor
        print(f"Fill color for cell {cell.coordinate}: {color}")

    if not color:
        print(f"Fill color for cell {cell.coordinate}: {color}")
        return 'FFFFFFFF'

    tint = getattr(color, 'tint', None)
    
    # RGB type color
    if getattr(color, 'rgb', None):
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb = rgb_from_argb(rgb_val)
            if tint:  # apply tint
                rgb = apply_tint(rgb, tint)
            return rgb_to_hex(rgb)
        elif len(rgb_val) == 6:
            rgb = tuple(int(rgb_val[i:i+2], 16) for i in (0, 2, 4))
            if tint:
                rgb = apply_tint(rgb, tint)
            return rgb_to_hex(rgb)

    # Indexed colors
    if getattr(color, 'indexed', None) is not None:
        try:
            idx = int(color.indexed)
            rgb_val = COLOR_INDEX[idx] if 0 <= idx < len(COLOR_INDEX) else None
            if rgb_val:
                rgb = rgb_from_argb(rgb_val)
                if tint:
                    rgb = apply_tint(rgb, tint)
                return rgb_to_hex(rgb)
        except Exception:
            pass

    # Theme colors - fallback or implement theme mapping here (not covered)
    print(f"Fill color for cell {cell.coordinate}: {color}")
    return 'FFFFFFFF'  # fallback white


def get_font_rgb(cell):
    color = None
    try:
        color = cell.font.color
    except AttributeError:
        return "FF000000"  # default black if no font or color attribute

    if color is None:
        return "FF000000"

    tint = getattr(color, 'tint', None)

    if hasattr(color, 'rgb') and color.rgb:
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb = rgb_from_argb(rgb_val)
            if tint:
                rgb = apply_tint(rgb, tint)
            return rgb_to_hex(rgb)
        elif len(rgb_val) == 6:
            rgb = tuple(int(rgb_val[i:i+2], 16) for i in (0, 2, 4))
            if tint:
                rgb = apply_tint(rgb, tint)
            return rgb_to_hex(rgb)

    if hasattr(color, 'indexed') and color.indexed is not None:
        try:
            idx = int(color.indexed)
            rgb_val = COLOR_INDEX[idx] if 0 <= idx < len(COLOR_INDEX) else None
            if rgb_val:
                rgb = rgb_from_argb(rgb_val)
                if tint:
                    rgb = apply_tint(rgb, tint)
                return rgb_to_hex(rgb)
        except Exception:
            pass

    # Theme colors fallback to black
    return "FF000000"

def has_strikethrough(cell):
    """
    Detects if ANY part of the cell text has strikethrough formatting.
    Checks rich-text runs first, falls back to cell.font.strike.
    """
    try:
        c_elem = cell._element

        # Look for rich-text runs
        runs = c_elem.findall(".//m:r", NS)
        for r in runs:
            rPr = r.find("m:rPr", NS)
            if rPr is None:
                continue
            strike_el = rPr.find("m:strike", NS)
            if strike_el is not None:
                val = strike_el.get("val", "1")
                if val in ("1", "true", None):
                    return True

        # Fallback: whole-cell font
        return bool(cell.font.strike)
    except Exception:
        return bool(cell.font.strike)


def extract_style(cell, style_cache):
    """
    Extracts and caches style info for a cell.
    Returns a dict with font/fill/alignment details.
    """
    if not cell.has_style:
        return None

    # Build a key to cache repeated styles
    key = (
        cell.font.bold, cell.font.italic, cell.font.underline,
        # getattr(cell.font.color, "rgb", None),
        # getattr(cell.fill.fgColor, "rgb", None),
        get_font_rgb(cell),
        get_fill_color(cell),
        has_strikethrough(cell),
        cell.number_format,
        cell.alignment.horizontal,
        cell.alignment.vertical
    )

    if key not in style_cache:
        style_cache[key] = {
            "font_bold": cell.font.bold,
            "font_italic": cell.font.italic,
            "font_underline": cell.font.underline,
            # "font_color": getattr(cell.font.color, "rgb", None),
            # "fill_color": getattr(cell.fill.fgColor, "rgb", None),
            "font_color": get_font_rgb(cell),
            "fill_color": get_fill_color(cell),
            "font_strike": has_strikethrough(cell),
            "number_format": cell.number_format,
            "alignment_horizontal": cell.alignment.horizontal,
            "alignment_vertical": cell.alignment.vertical,
        }

    return style_cache[key]


def expand_merged_cells(ws):
    """
    Expand merged cells into individual cells by copying
    the top-left cell's value and style into all merged cells.
    """
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_cell = ws.cell(row=min_row, column=min_col)
        top_left_value = top_left_cell.value

        # ✅ copy() to make independent style objects
        top_font = copy(top_left_cell.font)
        top_fill = copy(top_left_cell.fill)
        top_alignment = copy(top_left_cell.alignment)

        ws.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                c = ws.cell(row=row, column=col, value=top_left_value)
                c.font = copy(top_font)
                c.fill = copy(top_fill)
                c.alignment = copy(top_alignment)

def filter_latest_comments(text, max_comments=2):
    """
    Extract only the latest max_comments dated comments from a multi-line Comments field.
    Interprets dates in DD/MM format; sorts by full date assumed for current year.
    """
    # if pd.isna(text) or text in (None, ""): --> Pandas dependency removed
    #     return text
    if text is None or str(text).strip() == "" or str(text).lower() == "nan":
        return text

    comments = str(text).split("\n")

    comment_dates = []
    for c in comments:
        # Find all DD/MM date strings in comment
        dates = re.findall(r"\b(\d{2}/\d{2})\b", c)
        if dates:
            # Take the last date in the comment as its date
            last_date_str = dates[-1]
            try:
                # Parse DD/MM to datetime (day, month order)
                day, month = map(int, last_date_str.split("/"))
                # date_obj = datetime(datetime.now().year, month, day) --> handled below for invalid dates
                # Handle invalid dates (e.g., 31/02) by assigning minimal date
                current_year = datetime.now().year
                try:
                  date_obj = datetime(current_year, month, day)
                except ValueError:
                  date_obj = datetime.min
            except Exception:
                date_obj = datetime.min  # On parse error, assign minimal date
            comment_dates.append((c, date_obj))
        else:
            # Comments without dates get minimal date to be sorted last
            comment_dates.append((c, datetime.min))

    # Sort descending by date
    comment_dates.sort(key=lambda x: x[1], reverse=True)

    selected_comments = []
    seen_comments = set()
    for comment, _date in comment_dates:
        if comment not in seen_comments:
            selected_comments.append(comment)
            seen_comments.add(comment)
        if len(selected_comments) >= max_comments:
            break

    return "\n".join(selected_comments)

def extract_strike_segments(cell):
    """
    Because your Excel file contains NO real strikethrough formatting,
    we infer strike based on the rule:
    ➤ All values except the LAST one are struck.
    """
    if cell is None or cell.value is None:
        return []

    # Split multiline values
    lines = [line.strip() for line in str(cell.value).split("\n") if line.strip()]

    if not lines:
        return []

    segments = []

    for i, line in enumerate(lines):
        # All but last entry → struck
        strike = (i != len(lines) - 1)
        segments.append({
            "text": line,
            "strike": strike
        })

    return segments

def build_row_data(row, final_headers, keep_indices, style_cache, include_styles):
    """Builds a dictionary of cell values and styles for a row."""
    row_data = {}
    for pos in range(len(final_headers)):
        header = final_headers[pos]
        try:
            cell = row[keep_indices[pos]]
            value = cell.value
        except IndexError:
            cell, value = None, None

        if isinstance(value, datetime):
            value = value.date()
        if header.lower() == "comments":
            value = filter_latest_comments(value, max_comments=2)

        row_data[header] = {
            "value": value,
            "strike_segments": extract_strike_segments(cell) if cell is not None else None,
            "style": extract_style(cell, style_cache) if (include_styles and cell is not None) else None
        }
    return row_data

def is_section_title_row(row):
    """
    Robust detection of section title rows.

    A row is considered a section title if:
    - It has only one non-empty cell with meaningful text (len >= 4)
      (style cues like bold/large/fill increase confidence but are not mandatory)
    OR
    - It has multiple non-empty cells, all with the same text (after merge expansion),
      and that text is meaningful (len >= 4)
    """

    # Collect non-empty cells
    non_empty_cells = [c for c in row if c.value not in (None, "")]
    if not non_empty_cells:
        return False

    # --- Case 1: One non-empty cell ---
    if len(non_empty_cells) == 1:
        c = non_empty_cells[0]
        text = str(c.value).strip()
        if len(text) < 4:
            return False
        # Style is optional, just boosts confidence
        font = getattr(c, "font", None)
        fill = getattr(c, "fill", None)
        bold = bool(font and font.bold)
        big = bool(font and getattr(font, "sz", None) and font.sz == 11)
        filled = bool(fill and fill.fgColor and fill.fgColor.type == "rgb"
                      and fill.fgColor.rgb and fill.fgColor.rgb.upper() not in ("FFFFFFFF", "FF000000"))
        # Even if no style, allow it as section title
        return True

    # --- Case 2: Multiple identical non-empty cells (from merged expansion) ---
    # unique_vals = {str(c.value).strip() for c in non_empty_cells if str(c.value).strip()}
    # if len(unique_vals) == 1:
    #     text = next(iter(unique_vals))
    #     if len(text) >= 4 and len(non_empty_cells) >= 2:
    #         return True
    def normalize(s):
        return re.sub(r'\s+', ' ', str(s).strip())

    unique_vals = {normalize(c.value) for c in non_empty_cells if c.value is not None}

    if len(non_empty_cells) > 1:
        print(f"Row {row[0].row}: unique_vals = {[repr(normalize(c.value)) for c in non_empty_cells]}")

    if len(unique_vals) == 1:
        text = next(iter(unique_vals))
        if len(text) >= 4 and len(non_empty_cells) >= 2:
            return True

    return False

def is_true_header_row(values):
    # Normalize and check for key phrases
    header_phrases = {"s.no", "description/details"}
    norm = [str(v).strip().lower() for v in values if v]
    return any(phrase in norm for phrase in header_phrases)

def is_row_skippable(row, keep_indices=None):
    """
    Return True if all cells except the first one are empty.
    - row: list of openpyxl Cell objects
    - keep_indices: optional, list of column indices to consider
    """
    if keep_indices is None:
        # Default: consider all columns
        cells_to_check = row[1:]  # skip first cell
    else:
        # Only check columns from keep_indices, skipping first
        if len(keep_indices) <= 1:
            return False  # nothing to check
        cells_to_check = [row[idx] for idx in keep_indices[1:]]

    for c in cells_to_check:
        v = c.value
        if v is not None and str(v).strip() != "":
            return False
    return True


def is_phase_true_header(values):
    """Detects the real Phase column header row (Services, Days Used, etc)."""
    norm = [str(v).strip().lower() for v in values if v]
    keywords = {"services", "days used", "cost (excl vat)", "charges (excl. vat)", "days remaining", "balance"}
    match_count = sum(1 for v in norm if any(kw in v for kw in keywords))
    return match_count >= 3

def process_sheet(ws, output_folder, include_styles=True, stream=True, debug=False, prefix=""):
    sheet_name = ws.title
    style_cache = {}

    # Expand merged cells first
    expand_merged_cells(ws)

    tables = []
    current_table = None
    keep_indices = []
    final_headers = []
    pending_section_titles = []

    for i in range(1, ws.max_row + 1):
        # Skip hidden rows
        if ws.row_dimensions[i].hidden:
            continue
        row = ws[i]
        values = [str(c.value).strip() if c.value is not None else "" for c in row]

        # Skip empty row for table termination
        if all(v == "" for v in values):
            current_table = None
            continue

        # Detect section title row
        if is_section_title_row(row):
            section_cell = next(c for c in row if c.value)
            section_entry = {
                "section_title": str(section_cell.value).strip(),
                "style": extract_style(section_cell, style_cache),
                "row_idx": section_cell.row
            }
            if current_table:
                current_table["rows"].append(section_entry)
            else:
                pending_section_titles.append(section_entry)
            continue

        # Detect table header
        if current_table is None:
            non_empty_values = [v for v in values if v != ""]
            string_like_count = sum(1 for v in non_empty_values if isinstance(v, str) and len(v) <= 50)
            if len(non_empty_values) >= 2 and string_like_count >= 1:
                # Check if this is a "false" header row (e.g. merged Phase labels)
                if sheet_name.strip().lower() == "phase" and not is_phase_true_header(values):
                    # Peek next row: if it looks like a real header, skip this one
                    if i + 1 <= ws.max_row:
                        next_row = ws[i+1]
                        next_vals = [str(c.value).strip() if c.value is not None else "" for c in next_row]
                        if is_phase_true_header(next_vals):
                            continue

                keep_indices = [idx for idx, h in enumerate(values) if h != ""]
                final_headers = [values[idx] for idx in keep_indices]
                # Drop "Comments" column for Assumption sheet
                if sheet_name.strip().lower() == "assumptions":
                    filtered_indices = []
                    filtered_headers = []
                    for idx, header in zip(keep_indices, final_headers):
                        if str(header).strip().lower() != "comments":
                            filtered_indices.append(idx)
                            filtered_headers.append(header)
                    keep_indices = filtered_indices
                    final_headers = filtered_headers
                    print(f"Dropped 'Comments' column from Assumption sheet")

                current_table = {
                    "header_row": i,
                    "headers": final_headers,
                    "rows": pending_section_titles.copy(),
                }
                pending_section_titles.clear()
                tables.append(current_table)
                continue

        # Normal row inside table
        if current_table:
            if is_row_skippable(row, keep_indices):
                continue
            row_data = build_row_data(row, current_table["headers"], keep_indices, style_cache, include_styles)
            current_table["rows"].append(row_data)

    # Save JSON
    os.makedirs(output_folder, exist_ok=True)
    filename = f"{prefix}=={sheet_name}.json" if prefix else f"{sheet_name}.json"
    json_path = os.path.join(output_folder, filename)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(tables, f, indent=2, default=str)

    print(f"Saved JSON for sheet '{sheet_name}' → {json_path}, {len(tables)} table(s) detected.")
    return json_path


def get_project_start_date(sheets_data):
    for sheet in sheets_data:
        for table in sheet.get("tables", []):
            if "key milestone" in table.get("section_title", "").lower():
                # Find "Start Date" column
                start_date_col = None
                for h in table.get("headers", []):
                    if h.strip().lower() in ["start date", "startdate", "start_date"]:
                        start_date_col = h
                        break

                if start_date_col:
                    # Take first non-empty Start Date
                    for row in table.get("rows", []):
                        val = row.get(start_date_col, {}).get("value")
                        if val not in (None, ""):
                            # Format if datetime
                            if isinstance(val, datetime):
                                return val.strftime("%d-%b-%Y")
                            return str(val)
    return None


def read_visible_excel_sheets_to_json(file_path, output_folder="extracted_json", include_styles=True, stream=True, sheets=None, parallel=False):
    """
    Convert Excel sheets to JSON (scalable).
    - include_styles: whether to extract formatting info
    - stream: write JSON row by row (low memory)
    - sheets: list of sheet names to include (None = all visible)
    - parallel: process multiple sheets in parallel
    """
    os.makedirs(output_folder, exist_ok=True)

    # Use read_only if not extracting styles (faster for large files)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Select sheets
    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    visible_sheets = [ws for ws in visible_sheets if ws.title.lower() != "legend"]  # Exclude 'Legend'
    if sheets:
        visible_sheets = [ws for ws in visible_sheets if ws.title in sheets]

    saved_files = []

    file_prefix = os.path.splitext(os.path.basename(file_path))[0]

    if parallel:
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(process_sheet, ws, output_folder, include_styles, stream, prefix=file_prefix) for ws in visible_sheets]
            saved_files = [f.result() for f in futures if f.result()]
    else:
        for ws in visible_sheets:
            result = process_sheet(ws, output_folder, include_styles, stream, prefix=file_prefix)
            if result:
                saved_files.append(result)

    return saved_files